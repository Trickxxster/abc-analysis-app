import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="📊 Расчёт заказа и ABC-анализ", layout="wide")
st.title("📊 Автоматизация расчёта заказа и ABC-анализ (с ручной корректировкой)")

# ------------------------------------------------------------
# Жёсткий порядок городов
# ------------------------------------------------------------
CITY_ORDER = [
    "Абакан", "Архангельск", "Барнаул", "Иркутск", "Кемерово",
    "Красноярск", "Новокузнецк", "Омск", "Томск", "Хабаровск", "Чита"
]

# ------------------------------------------------------------
# Вспомогательные функции
# ------------------------------------------------------------
def safe_convert_to_float(series: pd.Series) -> pd.Series:
    if series.dtype == object:
        series = series.astype(str).str.replace(r'\s+', '', regex=True).str.replace(',', '.')
    return pd.to_numeric(series, errors='coerce').fillna(0)

def find_class_columns(df: pd.DataFrame) -> list:
    class_set = {'A', 'B', 'C', 'a', 'b', 'c'}
    class_cols = []
    for i in range(2, len(df.columns)):
        col_data = df.iloc[:, i]
        found = False
        for val in col_data:
            if pd.notna(val):
                s = str(val).strip()
                if s in class_set:
                    found = True
                    break
        if found:
            class_cols.append(i)
    return class_cols

def recalculate_all(df_raw, class_indices, target_days, days_in_report, manual_overrides):
    city_data = {}
    for idx, city in enumerate(CITY_ORDER):
        if idx >= len(class_indices):
            break
        start_col = class_indices[idx]
        block = df_raw.iloc[:, start_col:start_col+10]
        if block.isnull().all().all():
            continue

        col_sales_period = start_col + 1   # Количество (продано за период)
        col_ost = start_col + 5            # Итоговый остаток
        col_in_transit = start_col + 4     # Товары в пути

        city_df = df_raw[['Товар', col_sales_period, col_ost, col_in_transit]].copy()
        city_df.columns = ['Товар', 'Продано за период', 'Остаток', 'В пути']

        for col in ['Продано за период', 'Остаток', 'В пути']:
            city_df[col] = safe_convert_to_float(city_df[col])

        daily_sales = city_df['Продано за период'] / days_in_report
        city_df['Среднедневные продажи'] = daily_sales.round(4)

        need = (daily_sales * target_days) - city_df['Остаток'] - city_df['В пути']
        raw_need = np.ceil(np.maximum(need, 0)).astype(int)

        order_col = []
        for idx_row, product in enumerate(city_df['Товар']):
            key = (city, product)
            if key in manual_overrides and pd.notna(manual_overrides[key]):
                order_col.append(int(manual_overrides[key]))
            else:
                order_col.append(int(raw_need[idx_row]))
        city_df['К заказу'] = order_col

        # Индикатор дефицита (остаток < среднедневные * 3)
        def deficit_flag(row):
            if row['Среднедневные продажи'] == 0:
                return 'Норма'
            if row['Остаток'] < row['Среднедневные продажи'] * 3:
                return '⚠️ Дефицит'
            else:
                return 'Норма'
        city_df['Дефицит'] = city_df.apply(deficit_flag, axis=1)

        # ABC-анализ
        total_sales = city_df['Продано за период'].sum()
        if total_sales == 0:
            city_df['Категория ABC'] = 'C'
        else:
            sorted_df = city_df.sort_values('Продано за период', ascending=False).copy()
            sorted_df['cum_sales'] = sorted_df['Продано за период'].cumsum()
            sorted_df['cum_perc'] = sorted_df['cum_sales'] / total_sales

            def assign_abc(row):
                if row['Продано за период'] == 0:
                    return 'C'
                if row['cum_perc'] <= 0.8:
                    return 'A'
                elif row['cum_perc'] <= 0.95:
                    return 'B'
                else:
                    return 'C'

            sorted_df['Категория ABC'] = sorted_df.apply(assign_abc, axis=1)
            city_df = city_df.merge(sorted_df[['Товар', 'Категория ABC']], on='Товар', how='left')

        city_df.set_index('Товар', inplace=True)
        city_df = city_df[['Остаток', 'Продано за период', 'Среднедневные продажи', 'В пути',
                           'К заказу', 'Дефицит', 'Категория ABC']]
        city_data[city] = city_df
    return city_data

def build_matrix(city_data):
    all_products = sorted({p for cdf in city_data.values() for p in cdf.index})
    matrix = pd.DataFrame(index=all_products, columns=list(city_data.keys()))
    for city, cdf in city_data.items():
        matrix[city] = matrix.index.map(lambda p: cdf.loc[p, 'К заказу'] if p in cdf.index else 0)
    matrix['Итого'] = matrix.sum(axis=1)
    return matrix

# ------------------------------------------------------------
# Форматирование Excel (с дефицитом)
# ------------------------------------------------------------
def apply_excel_formatting(writer, city_data):
    wb = writer.book
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 1. Шапка – голубая
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        header_font = Font(bold=True)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        # 2. Чередование строк
        light_gray = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            fill = light_gray if row % 2 == 0 else white
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

        # 3. Жирный шрифт для заказа > 0 (в сводной – Итого > 0)
        col_order = None
        col_total = None
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val == "К заказу":
                col_order = col
            elif val == "Итого":
                col_total = col
        if col_order is not None:
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col_order).value
                # Если это формула – извлекаем результат
                if val is not None:
                    try:
                        val = float(val)
                    except:
                        val = 0
                    if val > 0:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).font = Font(bold=True)
        elif col_total is not None:
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col_total).value
                if val is not None:
                    try:
                        val = float(val)
                    except:
                        val = 0
                    if val > 0:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).font = Font(bold=True)

        # 4. Красный фон для строк с дефицитом (только на листах городов)
        if sheet_name in city_data:
            col_def = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "Дефицит":
                    col_def = col
                    break
            if col_def is not None:
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=col_def).value == "⚠️ Дефицит":
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = red_fill

# ------------------------------------------------------------
# Генерация Excel с формулами в сводной матрице
# ------------------------------------------------------------
def generate_excel(matrix, city_data):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. Записываем листы городов и запоминаем позиции
        city_order_col = {}
        city_product_rows = {}
        for city, cdf in city_data.items():
            sheet_name = str(city)[:31]
            df_city = cdf.reset_index()
            df_city.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            # Находим столбец "К заказу"
            col_order = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "К заказу":
                    col_order = col
                    break
            if col_order is None:
                continue
            city_order_col[city] = col_order
            # Запоминаем строки товаров
            product_rows = {}
            for row in range(2, ws.max_row + 1):
                product_name = ws.cell(row=row, column=1).value
                if product_name is not None:
                    product_rows[product_name] = row
            city_product_rows[city] = product_rows

        # 2. Создаём сводную матрицу с формулами
        matrix_sheet = writer.book.create_sheet("Сводная матрица")
        headers = ["Товар"] + list(city_data.keys()) + ["Итого"]
        for col_idx, header in enumerate(headers, start=1):
            matrix_sheet.cell(row=1, column=col_idx, value=header)

        for row_idx, product in enumerate(matrix.index, start=2):
            matrix_sheet.cell(row=row_idx, column=1, value=product)
            for col_idx, city in enumerate(city_data.keys(), start=2):
                if city in city_product_rows and product in city_product_rows[city]:
                    row_num = city_product_rows[city][product]
                    col_letter = get_column_letter(city_order_col[city])
                    formula = f"='{city}'!{col_letter}{row_num}"
                    matrix_sheet.cell(row=row_idx, column=col_idx, value=formula)
                else:
                    matrix_sheet.cell(row=row_idx, column=col_idx, value=0)
            # Итого – сумма
            col_start = 2
            col_end = 2 + len(city_data) - 1
            sum_formula = f"=SUM({get_column_letter(col_start)}{row_idx}:{get_column_letter(col_end)}{row_idx})"
            matrix_sheet.cell(row=row_idx, column=col_end + 1, value=sum_formula)

        # 3. Стилизация
        apply_excel_formatting(writer, city_data)
    output.seek(0)
    return output

# ------------------------------------------------------------
# Инициализация состояния
# ------------------------------------------------------------
for key in ['manual_overrides', 'city_data', 'matrix', 'target_days', 'days_in_report',
            'df_raw', 'class_indices', 'uploaded_file']:
    if key not in st.session_state:
        if key == 'target_days':
            st.session_state[key] = 14
        elif key == 'days_in_report':
            st.session_state[key] = 30
        else:
            st.session_state[key] = None

# ------------------------------------------------------------
# Боковая панель
# ------------------------------------------------------------
st.sidebar.header("⚙️ Загрузка данных и настройки")
uploaded_file = st.sidebar.file_uploader("Загрузите Excel-файл с данными (без заголовков)", type=["xlsx", "xls"])
days_in_report = st.sidebar.number_input(
    "Количество дней в выгрузке продаж (из 1С)",
    min_value=1, max_value=365, value=st.session_state.days_in_report, step=1,
    help="Используется для расчёта среднедневных продаж."
)
target_days = st.sidebar.slider(
    "Целевой период обеспечения (дней)",
    min_value=1, max_value=90, value=st.session_state.target_days, step=1
)
reset_overrides = st.sidebar.button("🔄 Сбросить все ручные правки")

st.sidebar.markdown("---")
st.sidebar.info(
    "**Формула расчёта:**\n\n"
    "1. Продано за период – из столбца «Количество».\n"
    "2. Среднедневные продажи = Продано за период / Дни выгрузки.\n"
    "3. Потребность = (Среднедневные × Целевой период) – Остаток – В пути.\n"
    "4. Если потребность < 0 → 0.\n"
    "5. Округление до целого вверх.\n\n"
    "🔴 **Дефицит:** остаток < среднедневные × 3 дня.\n\n"
    "💡 Редактируйте **«К заказу»** во вкладке «Детализация»."
)

# ------------------------------------------------------------
# Основная логика
# ------------------------------------------------------------
if uploaded_file is not None:
    # Загрузка нового файла
    if st.session_state.uploaded_file != uploaded_file:
        df_raw = pd.read_excel(uploaded_file, header=None)
        product_col = df_raw[0].astype(str).fillna('')
        char_col = df_raw[1].astype(str).fillna('')
        df_raw['Товар'] = (product_col + ' ' + char_col).str.strip()
        st.session_state.df_raw = df_raw
        st.session_state.class_indices = find_class_columns(df_raw)
        st.session_state.uploaded_file = uploaded_file
        st.session_state.manual_overrides = {}
        st.session_state.target_days = target_days
        st.session_state.days_in_report = days_in_report
        # Пересчёт и принудительный rerun
        city_data = recalculate_all(
            st.session_state.df_raw,
            st.session_state.class_indices,
            st.session_state.target_days,
            st.session_state.days_in_report,
            st.session_state.manual_overrides
        )
        matrix = build_matrix(city_data)
        st.session_state.city_data = city_data
        st.session_state.matrix = matrix
        st.rerun()

    # Изменение слайдера или дней – сброс правок и пересчёт
    if (target_days != st.session_state.target_days or 
        days_in_report != st.session_state.days_in_report):
        st.session_state.manual_overrides = {}
        st.session_state.target_days = target_days
        st.session_state.days_in_report = days_in_report
        city_data = recalculate_all(
            st.session_state.df_raw,
            st.session_state.class_indices,
            st.session_state.target_days,
            st.session_state.days_in_report,
            st.session_state.manual_overrides
        )
        matrix = build_matrix(city_data)
        st.session_state.city_data = city_data
        st.session_state.matrix = matrix
        st.rerun()

    # Кнопка сброса
    if reset_overrides:
        st.session_state.manual_overrides = {}
        city_data = recalculate_all(
            st.session_state.df_raw,
            st.session_state.class_indices,
            st.session_state.target_days,
            st.session_state.days_in_report,
            st.session_state.manual_overrides
        )
        matrix = build_matrix(city_data)
        st.session_state.city_data = city_data
        st.session_state.matrix = matrix
        st.rerun()

    # Если данные ещё не загружены (первый запуск)
    if st.session_state.city_data is None and st.session_state.df_raw is not None:
        city_data = recalculate_all(
            st.session_state.df_raw,
            st.session_state.class_indices,
            st.session_state.target_days,
            st.session_state.days_in_report,
            st.session_state.manual_overrides
        )
        matrix = build_matrix(city_data)
        st.session_state.city_data = city_data
        st.session_state.matrix = matrix
        st.rerun()

    if st.session_state.city_data is None:
        st.error("Не удалось обработать данные. Проверьте структуру файла.")
        st.stop()

    city_data = st.session_state.city_data
    matrix = st.session_state.matrix

    st.success(f"Успешно загружены данные по {len(city_data)} городам: {', '.join(city_data.keys())}")

    # Вкладки
    tab1, tab2, tab3 = st.tabs(['📋 Сводная матрица', '🏙️ Детализация по городам (с редактированием)', '📈 Интерактивные графики'])

    with tab1:
        st.subheader('Общая потребность в дозаказе по городам')
        st.dataframe(matrix, use_container_width=True)

        excel_data = generate_excel(matrix, city_data)
        st.download_button(
            label='📥 Скачать итоговый Excel-файл',
            data=excel_data,
            file_name='ABC_заказ.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    with tab2:
        st.subheader('Детализация по выбранному городу (редактируйте «К заказу»)')
        selected_city = st.selectbox('Выберите город', list(city_data.keys()))

        if selected_city:
            cdf = city_data[selected_city].reset_index()
            display_cols = ['Товар', 'Остаток', 'Продано за период', 'Среднедневные продажи',
                            'В пути', 'К заказу', 'Дефицит', 'Категория ABC']
            edit_df = cdf[display_cols].copy()

            edited_df = st.data_editor(
                edit_df,
                column_config={
                    "К заказу": st.column_config.NumberColumn(
                        "К заказу",
                        help="Введите нужное количество. Изменения сохраняются.",
                        min_value=0,
                        step=1
                    )
                },
                disabled=['Товар', 'Остаток', 'Продано за период', 'Среднедневные продажи',
                          'В пути', 'Дефицит', 'Категория ABC'],
                use_container_width=True,
                key=f"editor_{selected_city}"
            )

            # Обновляем manual_overrides по результатам редактирования
            changed = False
            for idx, row in edited_df.iterrows():
                product = row['Товар']
                new_val = row['К заказу']
                key = (selected_city, product)
                if pd.isna(new_val):
                    if key in st.session_state.manual_overrides:
                        del st.session_state.manual_overrides[key]
                        changed = True
                else:
                    if key not in st.session_state.manual_overrides or st.session_state.manual_overrides[key] != int(new_val):
                        st.session_state.manual_overrides[key] = int(new_val)
                        changed = True

            if changed:
                city_data_new = recalculate_all(
                    st.session_state.df_raw,
                    st.session_state.class_indices,
                    st.session_state.target_days,
                    st.session_state.days_in_report,
                    st.session_state.manual_overrides
                )
                matrix_new = build_matrix(city_data_new)
                st.session_state.city_data = city_data_new
                st.session_state.matrix = matrix_new
                city_data = city_data_new
                matrix = matrix_new
                st.rerun()

            total_order = cdf['К заказу'].sum() if 'К заказу' in cdf else 0
            st.metric(f'Всего единиц к заказу в г. {selected_city}', total_order)

    with tab3:
        st.subheader('Интерактивные графики продаж')
        
        st.markdown("#### Продажи выбранного товара по городам")
        all_products = sorted(matrix.index.tolist())
        selected_product = st.selectbox('Выберите товар', all_products, key='product_select')
        if selected_product:
            sales_data = {}
            for city, cdf in st.session_state.city_data.items():
                if selected_product in cdf.index:
                    sales_data[city] = cdf.loc[selected_product, 'Продано за период']
                else:
                    sales_data[city] = 0
            sales_df = pd.DataFrame(list(sales_data.items()), columns=['Город', 'Продажи, шт.'])
            fig1 = px.bar(
                sales_df, x='Город', y='Продажи, шт.',
                title=f'Продажи за период: {selected_product}',
                labels={'Продажи, шт.': 'Продано за период'},
                text_auto=True, color='Продажи, шт.',
                color_continuous_scale='Viridis'
            )
            fig1.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig1, use_container_width=True)
        
        st.markdown("---")
        st.markdown("#### Суммарные продажи по всем товарам (все города)")
        
        total_sales_per_product = {}
        for city, cdf in st.session_state.city_data.items():
            for product in cdf.index:
                total_sales_per_product[product] = total_sales_per_product.get(product, 0) + cdf.loc[product, 'Продано за период']
        
        sorted_products = sorted(total_sales_per_product.items(), key=lambda x: x[1], reverse=True)
        df_total = pd.DataFrame(sorted_products, columns=['Товар', 'Суммарные продажи, шт.'])
        
        fig2 = px.bar(
            df_total, x='Товар', y='Суммарные продажи, шт.',
            title='Суммарные продажи по всем товарам (все города)',
            labels={'Суммарные продажи, шт.': 'Продано за период'},
            text_auto=True, color='Суммарные продажи, шт.',
            color_continuous_scale='Plasma'
        )
        fig2.update_layout(xaxis_tickangle=-90)
        st.plotly_chart(fig2, use_container_width=True)

else:
    st.info('👈 Загрузите Excel-файл через боковую панель, чтобы начать работу.')
    st.markdown("""
    ### Ожидаемая структура данных (без заголовков)
    - **Столбец A** – номенклатура товара  
    - **Столбец B** – характеристика (модель, цвет и т.п.)  
    - Далее идут блоки по **10 столбцов** для каждого города в строгом порядке:  
      Абакан, Архангельск, Барнаул, Иркутск, Кемерово, Красноярск, Новокузнецк, Омск, Томск, Хабаровск, Чита  
    - Внутри блока: **Класс (A/B/C), Количество (продажи за период), Остаток, Свободный, Товары в пути, Итоговый остаток, Ср. продажи, Себестоимость, Остаток по себест., Расчёт к заказу**  
    - Программа автоматически найдёт начала блоков по столбцам с классами.
    """)
